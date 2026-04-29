import os
import sqlite3
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import re
import json
import logging
import io
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass
from pathlib import Path

# AI/ML imports (optional) - allow the app to start even if heavy ML libs are missing
import google.generativeai as genai

# Import sentence-transformers and faiss for vector similarity
try:
    from sentence_transformers import SentenceTransformer
    HAS_SENTENCE_TRANSFORMERS = True
except ImportError:
    SentenceTransformer = None
    HAS_SENTENCE_TRANSFORMERS = False

try:
    import faiss
    HAS_FAISS = True
except ImportError:
    faiss = None
    HAS_FAISS = False

from dotenv import load_dotenv

# Web interface imports
from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import plotly.graph_objs as go
import plotly.utils
import plotly.io as pio
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import xlsxwriter
import requests  # For Ollama API calls

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/nlsql_system.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

@dataclass
class QueryResult:
    """Data class for query results"""
    success: bool
    data: Optional[pd.DataFrame]
    sql_query: str
    execution_time: float
    error_message: Optional[str] = None
    visualization_path: Optional[str] = None

class EntityRecognizer:
    """Module for recognizing entities in natural language queries"""
    
    def __init__(self):
        self.entity_patterns = {
            'table_names': r'\b(?:table|from|into|update|join)\s+(\w+)\b',
            'column_names': r'\b(?:select|where|order by|group by)\s+([a-zA-Z_]\w*(?:\s*,\s*[a-zA-Z_]\w*)*)',
            'numeric_values': r'\b\d+(?:\.\d+)?\b',
            'string_values': r"'([^']*)'|\"([^\"]*)\"",
            'operators': r'\b(?:=|!=|<|>|<=|>=|like|in|between)\b',
            'aggregations': r'\b(?:sum|avg|count|max|min|group by)\b',
        }
        
    def extract_entities(self, query: str) -> Dict[str, List[str]]:
        """Extract entities from natural language query"""
        entities = {}
        
        for entity_type, pattern in self.entity_patterns.items():
            matches = re.findall(pattern, query, re.IGNORECASE)
            entities[entity_type] = [match for match in matches if match]
            
        return entities
    
    def map_to_schema(self, entities: Dict, schema: Dict) -> Dict:
        """Map extracted entities to database schema"""
        mapped_entities = {}
        
        # Map table names
        if 'table_names' in entities:
            mapped_entities['tables'] = []
            for table in entities['table_names']:
                if table.lower() in [t.lower() for t in schema.keys()]:
                    mapped_entities['tables'].append(table)
                    
        return mapped_entities

class SimilarityIndex:
    """Module for finding similar queries using vector embeddings with persistence"""
    
    def __init__(self, model_name: str = 'paraphrase-multilingual-MiniLM-L12-v2', 
                 index_path: str = 'data/query_cache.faiss',
                 metadata_path: str = 'data/query_metadata.json'):
        if not HAS_SENTENCE_TRANSFORMERS:
            raise RuntimeError('sentence-transformers is not installed; similarity features disabled')

        self.model = SentenceTransformer(model_name)
        self.index = None
        self.query_cache = {}
        self.embeddings = []
        self.queries = []
        self.index_path = index_path
        self.metadata_path = metadata_path
        
        # Create data directory if it doesn't exist
        Path('data').mkdir(exist_ok=True)
        
        # Load existing index and metadata if available
        self._load_from_disk()
        
    def _normalize_for_embedding(self, query: str) -> str:
        """
        Normalize a query for embedding by replacing numbers with placeholders.
        This allows matching queries that differ only in numeric values.
        
        Examples:
            "Show employees hired in 2023" -> "Show employees hired in <NUM>"
            "2023లో చేరిన ఉద్యోగులను చూపించు" -> "<NUM>లో చేరిన ఉద్యోగులను చూపించు"
        """
        # Replace all numbers (including decimals) with a placeholder
        # Use \d+(?:\.\d+)? to match integers and decimals without word boundary
        normalized = re.sub(r'\d+(?:\.\d+)?', '<NUM>', query)
        return normalized
        
    def _load_from_disk(self):
        """Load FAISS index and query metadata from disk"""
        try:
            if os.path.exists(self.index_path) and os.path.exists(self.metadata_path):
                # Load FAISS index
                self.index = faiss.read_index(self.index_path)
                
                # Load metadata
                with open(self.metadata_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.queries = data.get('queries', [])
                    self.query_cache = data.get('query_cache', {})
                    
                logger.info(f"[OK] Loaded {len(self.queries)} cached queries from disk")
            else:
                logger.info("📝 No existing query cache found - will create new one")
        except Exception as e:
            logger.warning(f"Failed to load query cache: {e}")
            self.index = None
            self.queries = []
            self.query_cache = {}
    
    def _save_to_disk(self):
        """Save FAISS index and query metadata to disk for persistence"""
        try:
            if self.index is not None:
                # Save FAISS index
                faiss.write_index(self.index, self.index_path)
                
                # Save metadata
                metadata = {
                    'queries': self.queries,
                    'query_cache': self.query_cache,
                    'last_updated': datetime.now().isoformat()
                }
                
                with open(self.metadata_path, 'w', encoding='utf-8') as f:
                    json.dump(metadata, f, indent=2, ensure_ascii=False)
                
                logger.debug(f"[SAVE] Saved {len(self.queries)} queries to disk")
        except Exception as e:
            logger.error(f"Failed to save query cache: {e}")
        
    def add_query(self, query: str, sql: str, result_metadata: Dict):
        """Add a successful query to the similarity index and persist to disk"""
        # Use normalized query for embedding (numbers replaced with <NUM>)
        # This allows matching queries that differ only in numeric values
        normalized_query = self._normalize_for_embedding(query)
        embedding = self.model.encode([normalized_query])
        
        if self.index is None:
            dimension = embedding.shape[1]
            self.index = faiss.IndexFlatL2(dimension)
            
        self.index.add(embedding)
        self.embeddings.append(embedding[0])
        self.queries.append({
            'natural_query': query,  # Store original query for display
            'normalized_query': normalized_query,  # Store normalized for reference
            'sql_query': sql,
            'metadata': result_metadata,
            'timestamp': datetime.now().isoformat()
        })
        
        self.query_cache[query] = {
            'sql': sql,
            'metadata': result_metadata
        }
        
        # Persist to disk
        self._save_to_disk()
        
    def find_similar(self, query: str, k: int = 5, threshold: float = 0.70) -> List[Dict]:
        """
        Find similar queries in the index
        Args:
            query: The query to search for
            k: Number of nearest neighbors to search
            threshold: Minimum similarity score (0.0-1.0). Default 0.70 = 70% match
        """
        if self.index is None or len(self.queries) == 0:
            return []
        
        # Normalize query for embedding (replace numbers with <NUM>)
        # This allows matching queries that differ only in numeric values
        normalized_query = self._normalize_for_embedding(query)
        logger.debug(f"[FIND] Searching with normalized query: {normalized_query}")
        
        query_embedding = self.model.encode([normalized_query])
        distances, indices = self.index.search(query_embedding, min(k, len(self.queries)))
        
        similar_queries = []
        for i, (distance, idx) in enumerate(zip(distances[0], indices[0])):
            # FAISS L2 distance: lower is more similar
            # Convert to cosine-like similarity: 1/(1+distance)
            # For L2 distance, typical range: 0 (identical) to 2 (very different)
            # This gives similarity range: 1.0 (identical) to 0.33 (very different)
            similarity = 1 / (1 + distance)
            
            # Log for debugging
            logger.debug(f"Query match: distance={distance:.4f}, similarity={similarity:.4f}, threshold={threshold}")
            
            if similarity >= threshold:
                similar_queries.append({
                    'query': self.queries[idx],
                    'similarity': float(similarity),  # Convert to Python float for JSON
                    'distance': float(distance)
                })
                logger.info(f"[MATCH] Match found: {self.queries[idx]['natural_query'][:50]}... ({similarity*100:.1f}%)")
                
        if not similar_queries:
            logger.info(f"[NO-MATCH] No matches above {threshold*100:.0f}% threshold (best: {(1/(1+distances[0][0]))*100:.1f}%)")
                
        return similar_queries
    
    def swap_entities(self, original_query: str, new_query: str, original_sql: str) -> Dict:
        """
        Smart entity swapping: Replace entities from original query with entities from new query
        Handles numbers, strings, dates, and other entities intelligently
        Works with multilingual queries (English, Telugu, Hindi, etc.)
        
        Returns:
            dict with keys:
                - adapted_sql: The SQL with swapped entities
                - swapped: Whether any entities were swapped
                - structural_change: Whether the query structure differs significantly
                - message: Human-readable message
        """
        import re
        
        # Detect structural differences that can't be handled by simple swapping
        structural_keywords_orig = set(re.findall(r'\b(show|display|get|select|list|find|count|sum|avg|max|min|group by|order by|limit|join)\b', original_query.lower()))
        structural_keywords_new = set(re.findall(r'\b(show|display|get|select|list|find|count|sum|avg|max|min|group by|order by|limit|join)\b', new_query.lower()))
        
        # Check if there are column-specific requests
        column_pattern = r'\b(name|age|salary|department|email|phone|id|hire_date|experience)\b'
        orig_columns = set(re.findall(column_pattern, original_query.lower()))
        new_columns = set(re.findall(column_pattern, new_query.lower()))
        
        # Detect if SELECT columns might be different
        has_column_change = False
        if orig_columns != new_columns and len(new_columns) > 0:
            # User is explicitly asking for specific columns
            has_column_change = True
            logger.warning(f"[WARN] Column change detected: {orig_columns} -> {new_columns}")
        
        # Detect if aggregation changes (count, sum, avg, etc.)
        has_aggregation_change = bool(structural_keywords_orig - structural_keywords_new or structural_keywords_new - structural_keywords_orig)
        
        structural_change = has_column_change or has_aggregation_change
        
        # Extract numbers from both queries - improved regex for multilingual support
        # This pattern finds numbers even when attached to non-ASCII characters (Telugu, Hindi, etc.)
        original_numbers = re.findall(r'(\d+(?:\.\d+)?)', original_query)
        new_numbers = re.findall(r'(\d+(?:\.\d+)?)', new_query)
        
        # Extract quoted strings from both queries
        original_strings = re.findall(r"'([^']*)'|\"([^\"]*)\"", original_query)
        new_strings = re.findall(r"'([^']*)'|\"([^\"]*)\"", new_query)
        original_strings = [s[0] or s[1] for s in original_strings]
        new_strings = [s[0] or s[1] for s in new_strings]
        
        adapted_sql = original_sql
        swapped = False
        
        # Swap numbers in SQL
        if original_numbers and new_numbers:
            for orig_num, new_num in zip(original_numbers, new_numbers):
                if orig_num != new_num:
                    # Replace in SQL - handle both word-boundary and non-boundary cases
                    # First try with word boundaries
                    new_sql = re.sub(
                        r'\b' + re.escape(orig_num) + r'\b',
                        new_num,
                        adapted_sql,
                        count=1
                    )
                    if new_sql == adapted_sql:
                        # If no match with word boundaries, try matching the number in patterns like '2023%'
                        new_sql = re.sub(
                            re.escape(orig_num),
                            new_num,
                            adapted_sql,
                            count=1
                        )
                    adapted_sql = new_sql
                    swapped = True
                    logger.info(f"[SWAP] Entity swap: {orig_num} -> {new_num}")
        
        # Swap quoted strings in SQL
        if original_strings and new_strings:
            for orig_str, new_str in zip(original_strings, new_strings):
                if orig_str != new_str:
                    # Replace both single and double quoted versions
                    adapted_sql = adapted_sql.replace(f"'{orig_str}'", f"'{new_str}'")
                    adapted_sql = adapted_sql.replace(f'"{orig_str}"', f'"{new_str}"')
                    # Also handle UPPER() case-insensitive versions
                    adapted_sql = adapted_sql.replace(
                        f"UPPER('{orig_str}')", 
                        f"UPPER('{new_str}')"
                    )
                    swapped = True
                    logger.info(f"[SWAP] Entity swap: '{orig_str}' -> '{new_str}'")
        
        # Determine message
        if structural_change:
            message = "[WARN] Structural change detected - recommend generating new SQL"
        elif swapped:
            message = "[OK] Entities swapped successfully"
        else:
            message = "No entities to swap"
        
        return {
            'adapted_sql': adapted_sql,
            'swapped': swapped,
            'structural_change': structural_change,
            'message': message
        }

class VisualizationEngine:
    """Module for creating dynamic visualizations with both Matplotlib and Plotly"""
    
    def __init__(self, output_dir: str = "visualizations"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        plt.style.use('seaborn-v0_8')
        
    def create_plotly_visualization(self, df: pd.DataFrame, chart_type: str = None, 
                                  title: str = None) -> Dict:
        """Create interactive Plotly visualization"""
        if df.empty:
            return None
            
        if chart_type is None:
            chart_type = self.detect_chart_type(df)
            
        if chart_type == 'none':
            return None
            
        try:
            if chart_type == 'bar':
                fig_data = self._create_plotly_bar(df)
            elif chart_type == 'line':
                fig_data = self._create_plotly_line(df)
            elif chart_type == 'pie':
                fig_data = self._create_plotly_pie(df)
            elif chart_type == 'histogram':
                fig_data = self._create_plotly_histogram(df)
            elif chart_type == 'scatter':
                fig_data = self._create_plotly_scatter(df)
            else:
                fig_data = self._create_plotly_bar(df)  # Default
                
            layout = go.Layout(
                title=title or f'Data Visualization ({chart_type.title()} Chart)',
                xaxis=dict(title=df.columns[0] if len(df.columns) > 0 else 'X-axis'),
                yaxis=dict(title=df.columns[1] if len(df.columns) > 1 else 'Y-axis'),
                hovermode='closest',
                template='plotly_white'
            )
            
            # Convert Plotly graph objects into JSON-serializable dict using plotly.io
            fig_obj = {
                'data': fig_data,
                'layout': layout
            }
            try:
                fig_json = json.loads(pio.to_json(fig_obj))
                return fig_json
            except Exception as e:
                # Fallback: try using PlotlyJSONEncoder directly
                try:
                    fig_json_str = json.dumps(fig_obj, cls=plotly.utils.PlotlyJSONEncoder)
                    return json.loads(fig_json_str)
                except Exception as e2:
                    logger.error(f"Failed to serialize Plotly figure: {e2}")
                    return None
            
        except Exception as e:
            logger.error(f"Error creating Plotly visualization: {e}")
            return None
    
    def _create_plotly_bar(self, df: pd.DataFrame):
        """Create Plotly bar chart"""
        if len(df.columns) >= 2:
            x_col, y_col = df.columns[0], df.columns[1]
            return [go.Bar(x=df[x_col], y=df[y_col], name=y_col)]
        else:
            values = df.iloc[:, 0].value_counts()
            return [go.Bar(x=values.index, y=values.values, name=df.columns[0])]
            
    def _create_plotly_line(self, df: pd.DataFrame):
        """Create Plotly line chart"""
        if len(df.columns) >= 2:
            x_col, y_col = df.columns[0], df.columns[1]
            return [go.Scatter(x=df[x_col], y=df[y_col], mode='lines+markers', name=y_col)]
        else:
            return [go.Scatter(y=df.iloc[:, 0], mode='lines+markers', name=df.columns[0])]
            
    def _create_plotly_pie(self, df: pd.DataFrame):
        """Create Plotly pie chart"""
        if len(df.columns) >= 2:
            labels_col, values_col = df.columns[0], df.columns[1]
            return [go.Pie(labels=df[labels_col], values=df[values_col])]
        else:
            values = df.iloc[:, 0].value_counts()
            return [go.Pie(labels=values.index, values=values.values)]
            
    def _create_plotly_histogram(self, df: pd.DataFrame):
        """Create Plotly histogram"""
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 0:
            return [go.Histogram(x=df[numeric_cols[0]], name=numeric_cols[0])]
        else:
            return [go.Histogram(x=df.iloc[:, 0], name=df.columns[0])]
            
    def _create_plotly_scatter(self, df: pd.DataFrame):
        """Create Plotly scatter plot"""
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) >= 2:
            x_col, y_col = numeric_cols[0], numeric_cols[1]
            return [go.Scatter(x=df[x_col], y=df[y_col], mode='markers', name=f'{y_col} vs {x_col}')]
        else:
            return [go.Scatter(y=df.iloc[:, 0], mode='markers', name=df.columns[0])]
        
    def detect_chart_type(self, df: pd.DataFrame) -> str:
        """Auto-detect appropriate chart type based on data"""
        if df.empty:
            return 'none'
            
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        categorical_cols = df.select_dtypes(exclude=[np.number]).columns
        
        if len(numeric_cols) >= 2:
            return 'scatter'
        elif len(numeric_cols) == 1 and len(categorical_cols) == 1:
            if df[categorical_cols[0]].nunique() <= 10:
                return 'bar'
            else:
                return 'histogram'
        elif len(categorical_cols) == 1:
            return 'pie'
        elif len(numeric_cols) == 1:
            return 'histogram'
        else:
            return 'bar'
    
    def create_visualization(self, df: pd.DataFrame, chart_type: str = None, 
                           title: str = None) -> str:
        """Create visualization based on data and chart type"""
        if df.empty:
            logger.warning("Cannot create visualization: DataFrame is empty")
            return None
            
        if chart_type is None:
            chart_type = self.detect_chart_type(df)
            
        if chart_type == 'none':
            return None
            
        plt.figure(figsize=(12, 8))
        
        try:
            if chart_type == 'bar':
                self._create_bar_chart(df)
            elif chart_type == 'line':
                self._create_line_chart(df)
            elif chart_type == 'pie':
                self._create_pie_chart(df)
            elif chart_type == 'histogram':
                self._create_histogram(df)
            elif chart_type == 'scatter':
                self._create_scatter_plot(df)
            else:
                self._create_bar_chart(df)  # Default fallback
                
            if title:
                plt.title(title, fontsize=16, fontweight='bold')
                
            plt.tight_layout()
            
            # Generate unique filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{chart_type}_chart_{timestamp}.png"
            filepath = self.output_dir / filename
            
            plt.savefig(filepath, dpi=300, bbox_inches='tight')
            plt.close()
            
            logger.info(f"Visualization saved: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error creating visualization: {e}")
            plt.close()
            return None
    
    def _create_bar_chart(self, df: pd.DataFrame):
        """Create bar chart"""
        if len(df.columns) >= 2:
            x_col, y_col = df.columns[0], df.columns[1]
            df.plot(x=x_col, y=y_col, kind='bar', color='skyblue')
            plt.xticks(rotation=45)
        else:
            df.plot(kind='bar', color='skyblue')
            
    def _create_line_chart(self, df: pd.DataFrame):
        """Create line chart"""
        if len(df.columns) >= 2:
            x_col, y_col = df.columns[0], df.columns[1]
            df.plot(x=x_col, y=y_col, kind='line', marker='o')
        else:
            df.plot(kind='line', marker='o')
            
    def _create_pie_chart(self, df: pd.DataFrame):
        """Create pie chart"""
        if len(df.columns) >= 2:
            labels_col, values_col = df.columns[0], df.columns[1]
            plt.pie(df[values_col], labels=df[labels_col], autopct='%1.1f%%')
        else:
            df.iloc[:, 0].value_counts().plot(kind='pie', autopct='%1.1f%%')
            
    def _create_histogram(self, df: pd.DataFrame):
        """Create histogram"""
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 0:
            df[numeric_cols[0]].hist(bins=20, color='skyblue', alpha=0.7)
            plt.xlabel(numeric_cols[0])
            plt.ylabel('Frequency')
            
    def _create_scatter_plot(self, df: pd.DataFrame):
        """Create scatter plot"""
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) >= 2:
            x_col, y_col = numeric_cols[0], numeric_cols[1]
            plt.scatter(df[x_col], df[y_col], alpha=0.6, color='skyblue')
            plt.xlabel(x_col)
            plt.ylabel(y_col)

class AdvancedTextToSQLConverter:
    """Main class for the advanced text-to-SQL conversion system"""
    
    def __init__(self, db_path: str = None):
        # Load environment variables
        load_dotenv()
        
        # Initialize components
        self.db_path = db_path or os.getenv('DB_PATH', 'data/advanced_nlsql.db')
        self.entity_recognizer = EntityRecognizer()
        # Initialize similarity index only if dependencies are available
        if HAS_SENTENCE_TRANSFORMERS and HAS_FAISS:
            try:
                self.similarity_index = SimilarityIndex()
            except Exception as e:
                logger.warning(f"Similarity index disabled: {e}")
                self.similarity_index = None
        else:
            logger.info("Similarity features disabled (missing sentence-transformers/faiss)")
            self.similarity_index = None
        self.visualization_engine = VisualizationEngine()
        
        # Initialize AI models
        self._setup_ai_models()

        # Run a lightweight startup health-check for Gemini (non-fatal)
        try:
            self._startup_gemini_healthcheck()
        except Exception as e:
            logger.debug(f"Gemini startup health-check raised an exception: {e}")
        
        # Initialize database
        self._initialize_database()
        
        # Create logs directory
        Path('logs').mkdir(exist_ok=True)
        
        logger.info("Advanced Text-to-SQL Converter initialized successfully")
    
    def _setup_ai_models(self):
        """Setup AI models for query generation"""
        # Check if we should use Ollama (offline)
        self.use_ollama = os.getenv('USE_OLLAMA', 'false').lower() == 'true'
        self.ollama_base_url = os.getenv('OLLAMA_BASE_URL', 'http://localhost:11434')
        self.ollama_model = os.getenv('OLLAMA_MODEL', 'mistral:latest')
        
        # Initialize both models so we have fallback options
        ollama_available = False
        if self.use_ollama:
            # Test Ollama connection
            if self._test_ollama_connection():
                logger.info(f"Ollama model '{self.ollama_model}' configured successfully (OFFLINE)")
                ollama_available = True
            else:
                logger.warning("Ollama not available, will use Gemini AI")
                self.use_ollama = False
        
        # Always setup Gemini AI as fallback (even when Ollama is primary)
        try:
            api_key = os.getenv('GEMINI_API_KEY')
            if api_key and api_key != 'your_gemini_api_key_here':
                genai.configure(api_key=api_key)
                # Use a model name that is available in the current API (full resource name)
                # Prefer the Flash model which is available on most accounts
                try:
                    # Use the 2.0 flash model as requested
                    self.gemini_model = genai.GenerativeModel('models/gemini-2.0-flash')
                    logger.info("Gemini AI model configured successfully: models/gemini-2.0-flash (ONLINE)")
                except Exception:
                    # Fall back to the generic flash-latest alias only if 2.0 not available
                    self.gemini_model = genai.GenerativeModel('models/gemini-flash-latest')
                    logger.info("Gemini AI model configured successfully: models/gemini-flash-latest (ONLINE)")
            else:
                self.gemini_model = None
                if not ollama_available:
                    logger.warning("Gemini API key not configured and Ollama not available")
                
        except Exception as e:
            logger.error(f"Error setting up Gemini AI: {e}")
            self.gemini_model = None
            
    def _test_ollama_connection(self) -> bool:
        """Test if Ollama is running and model is available"""
        try:
            # Test if Ollama server is running
            response = requests.get(f"{self.ollama_base_url}/api/tags", timeout=5)
            if response.status_code == 200:
                # Check if our model is available
                models = response.json().get('models', [])
                available_models = [model.get('name', '') for model in models]
                
                if self.ollama_model in available_models:
                    logger.info(f"[OK] Ollama model '{self.ollama_model}' is available")
                    return True
                else:
                    logger.error(f"❌ Ollama model '{self.ollama_model}' not found")
                    logger.info(f"Available models: {available_models}")
                    return False
            return False
        except Exception as e:
            logger.error(f"Ollama connection test failed: {e}")
            return False

    def _startup_gemini_healthcheck(self):
        """Lightweight health-check for Gemini model connectivity and quota/permission status.

        This method performs a minimal, non-destructive generation to surface common issues
        (invalid key, quota, permission). It is non-fatal and logs actionable messages.
        """
        # Only run when Gemini is configured and Ollama is not preferred
        if getattr(self, 'use_ollama', False) or not getattr(self, 'gemini_model', None):
            return

        try:
            # Small/safe prompt and token limit to minimize cost
            prompt = "Return the single token: OK"
            try:
                response = self.gemini_model.generate_content(prompt, max_output_tokens=1)
            except TypeError:
                # Fallback if the client does not accept max_output_tokens arg
                response = self.gemini_model.generate_content(prompt)

            txt = getattr(response, 'text', None) or str(response)
            logger.info(f"Gemini health-check OK (sample): {repr(txt)[:80]}")
        except Exception as e:
            msg = str(e).lower()
            if 'quota' in msg or '429' in msg:
                logger.warning("Gemini health-check: quota or rate-limit detected. Check Google Cloud billing/quotas.")
            elif 'api key' in msg or 'invalid' in msg or 'unauthorized' in msg:
                logger.warning("Gemini health-check: API key invalid or unauthorized. Verify GEMINI_API_KEY/GEMINI_API_KEYS.")
            else:
                logger.warning(f"Gemini health-check failed: {e}")
    
    def _initialize_database(self):
        """Initialize SQLite database with enhanced schema"""
        try:
            # Create data directory if it doesn't exist
            Path(self.db_path).parent.mkdir(parents=True, exist_ok=True)
            
            # We'll open short-lived connections per request to avoid cross-thread issues
            # Ensure database file exists by touching it via a connection
            conn = sqlite3.connect(self.db_path, check_same_thread=False)
            cur = conn.cursor()
            cur.execute("PRAGMA foreign_keys = ON")
            conn.commit()
            conn.close()
            
            # Create enhanced sample schema if tables don't exist
            self._create_sample_schema()
            
            logger.info(f"Database initialized: {self.db_path}")
            
        except Exception as e:
            logger.error(f"Error initializing database: {e}")
            raise
            
    def _create_sample_schema(self):
        """Create sample database schema with data"""
        try:
            # Use a short-lived connection for schema creation
            conn = sqlite3.connect(self.db_path, check_same_thread=False)
            cur = conn.cursor()
            cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
            existing_tables = [row[0] for row in cur.fetchall()]
            
            # Create employees table if it doesn't exist
            if 'employees' not in existing_tables:
                cur.execute("""
                    CREATE TABLE employees (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT NOT NULL,
                        department TEXT,
                        salary REAL,
                        hire_date DATE,
                        email TEXT,
                        phone TEXT,
                        age INTEGER,
                        experience_years INTEGER
                    )
                """)
                
                # Insert sample employee data
                sample_employees = [
                    ('John Doe', 'Engineering', 75000, '2023-01-15', 'john.doe@company.com', '555-0101', 32, 5),
                    ('Jane Smith', 'Marketing', 65000, '2023-02-20', 'jane.smith@company.com', '555-0102', 28, 3),
                    ('Mike Johnson', 'Engineering', 80000, '2023-01-10', 'mike.johnson@company.com', '555-0103', 35, 8),
                    ('Sarah Wilson', 'HR', 60000, '2023-03-05', 'sarah.wilson@company.com', '555-0104', 30, 4),
                    ('David Brown', 'Finance', 70000, '2023-02-15', 'david.brown@company.com', '555-0105', 33, 6),
                    ('Lisa Garcia', 'Engineering', 78000, '2023-01-25', 'lisa.garcia@company.com', '555-0106', 29, 4),
                    ('Robert Lee', 'Marketing', 62000, '2023-03-10', 'robert.lee@company.com', '555-0107', 27, 2),
                    ('Emily Davis', 'HR', 58000, '2023-02-28', 'emily.davis@company.com', '555-0108', 26, 2)
                ]
                
                cur.executemany("""
                    INSERT INTO employees (name, department, salary, hire_date, email, phone, age, experience_years)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, sample_employees)
                
            # Create departments table if it doesn't exist
            if 'departments' not in existing_tables:
                self.cursor.execute("""
                    CREATE TABLE departments (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT NOT NULL,
                        budget REAL,
                        manager_name TEXT,
                        location TEXT,
                        established_date DATE
                    )
                """)
                
                # Insert sample department data
                sample_departments = [
                    ('Engineering', 500000, 'Mike Johnson', 'Building A', '2020-01-01'),
                    ('Marketing', 200000, 'Jane Smith', 'Building B', '2020-01-01'),
                    ('HR', 150000, 'Sarah Wilson', 'Building C', '2020-01-01'),
                    ('Finance', 300000, 'David Brown', 'Building B', '2020-01-01'),
                    ('Sales', 250000, 'Tom Anderson', 'Building A', '2020-06-01')
                ]
                
                cur.executemany("""
                    INSERT INTO departments (name, budget, manager_name, location, established_date)
                    VALUES (?, ?, ?, ?, ?)
                """, sample_departments)
                
            # Create projects table if it doesn't exist
            if 'projects' not in existing_tables:
                self.cursor.execute("""
                    CREATE TABLE projects (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT NOT NULL,
                        department TEXT,
                        budget REAL,
                        start_date DATE,
                        end_date DATE,
                        status TEXT
                    )
                """)
                
                # Insert sample project data
                sample_projects = [
                    ('Website Redesign', 'Marketing', 50000, '2024-01-01', '2024-06-30', 'In Progress'),
                    ('Mobile App Development', 'Engineering', 120000, '2024-02-01', '2024-12-31', 'In Progress'),
                    ('HR System Upgrade', 'HR', 30000, '2024-03-01', '2024-08-31', 'Planning'),
                    ('Financial Audit', 'Finance', 25000, '2024-01-15', '2024-04-15', 'Completed'),
                    ('Customer Portal', 'Engineering', 80000, '2024-04-01', '2024-10-31', 'Planning')
                ]
                
                cur.executemany("""
                    INSERT INTO projects (name, department, budget, start_date, end_date, status)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, sample_projects)

            conn.commit()
            conn.close()
            logger.info("Sample database schema created successfully")
            
        except Exception as e:
            logger.error(f"Error creating sample schema: {e}")
    
    def get_comprehensive_schema(self) -> Dict[str, Dict]:
        """Get detailed database schema information"""
        schema = {}
        
        try:
            # Create a new connection for thread safety
            conn = sqlite3.connect(self.db_path, check_same_thread=False)
            cursor = conn.cursor()
            
            # Get all tables
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [row[0] for row in cursor.fetchall()]
            
            for table in tables:
                # Get column information
                cursor.execute(f"PRAGMA table_info({table})")
                columns = cursor.fetchall()
                
                schema[table] = {
                    'columns': {},
                    'row_count': 0
                }
                
                # Process column information
                for col in columns:
                    col_name, col_type, not_null, default, pk = col[1], col[2], col[3], col[4], col[5]
                    schema[table]['columns'][col_name] = {
                        'type': col_type,
                        'not_null': bool(not_null),
                        'default': default,
                        'primary_key': bool(pk)
                    }
                
                # Get row count
                cursor.execute(f"SELECT COUNT(*) FROM {table}")
                schema[table]['row_count'] = cursor.fetchone()[0]
                
            conn.close()
            return schema
                
        except Exception as e:
            logger.error(f"Error getting schema: {e}")
            if 'conn' in locals():
                conn.close()
            return {}
    
    def generate_sql_with_ai(self, natural_query: str, schema: Dict) -> str:
        """Generate SQL using AI models (Ollama offline or Gemini online)"""
        
        # Try Ollama first if enabled
        if self.use_ollama:
            sql_query = self._generate_sql_with_ollama(natural_query, schema)
            if sql_query:
                return sql_query
            else:
                logger.warning("Ollama failed, trying Gemini as fallback")
        
        # Fallback to Gemini
        if not self.gemini_model:
            logger.error("No AI model available for SQL generation")
            return None
            
        return self._generate_sql_with_gemini(natural_query, schema)
    
    def _generate_sql_with_ollama(self, natural_query: str, schema: Dict) -> str:
        """Generate SQL using Ollama (offline)"""
        try:
            schema_info = self._format_schema_for_prompt(schema)
            
            # Very clear, directive prompt
            prompt = f"""You are a SQL code generator. Generate ONLY SQL code, no explanations.

DATABASE TABLES:
{schema_info}

USER REQUEST: {natural_query}

IMPORTANT INSTRUCTIONS:
1. If user says "show", "list", "find", "get" → Use SELECT query
2. If user says "create table" → Use CREATE TABLE
3. If user says "add", "insert" → Use INSERT INTO
4. If user mentions a number comparison (greater than, less than, >, <) → Use WHERE clause
5. If user says "average" or "avg" → Use SELECT AVG()
6. If user says "count" → Use SELECT COUNT()

EXAMPLES:
Request: "Show all employees with salary greater than 70000"
SQL: SELECT * FROM employees WHERE salary > 70000;

Request: "Create table students with roll and marks"
SQL: CREATE TABLE students (roll_number INTEGER PRIMARY KEY, marks INTEGER);

Request: "Find average salary by department"
SQL: SELECT department, AVG(salary) FROM employees GROUP BY department;

Now generate SQL for: {natural_query}

SQL:"""
            
            payload = {
                "model": self.ollama_model,
                "prompt": prompt,
                "stream": False,
                "options": {
                    "temperature": 0.2,
                    "top_p": 0.9,
                    "num_predict": 300  # Increased for CREATE/INSERT statements
                }
            }
            
            response = requests.post(
                f"{self.ollama_base_url}/api/generate", 
                json=payload, 
                timeout=120  # Increased timeout to 2 minutes
            )
            
            if response.status_code == 200:
                result = response.json()
                sql_query = result.get('response', '').strip()
                
                # Clean up the response
                sql_query = self._clean_sql_query(sql_query)
                
                logger.info(f"[OK] Generated SQL with Ollama ({self.ollama_model}): {sql_query}")
                return sql_query
            else:
                logger.error(f"Ollama API error: {response.status_code} - {response.text}")
                return None
                
        except requests.exceptions.Timeout:
            logger.error(f"Ollama request timed out after 120s - model may be too slow")
            return None
        except Exception as e:
            logger.error(f"Error generating SQL with Ollama: {e}")
            return None
    
    def _generate_sql_with_gemini(self, natural_query: str, schema: Dict) -> str:
        """Generate SQL using Gemini AI (online)"""
        # Implement retries/backoff for 429 (quota) errors and a safe fallback.
        max_attempts = 3
        backoff_base = 1.5

        schema_info = self._format_schema_for_prompt(schema)
        prompt = f"""You are an expert SQL query generator. Your task is to convert natural language requests into precise, executable SQLite queries.

Database Schema:
{schema_info}

Natural Language Query: "{natural_query}"

Critical Requirements:
1. Return ONLY valid SQLite SQL - no markdown, no explanations, no preamble
2. Use exact table and column names from the schema above
3. Implement ALL filters, conditions, and aggregations mentioned in the user query
4. For string comparisons, use case-insensitive matching: WHERE UPPER(column) = UPPER('value')
5. For numeric comparisons (greater than, less than), use >, <, >=, <= operators
6. For aggregations (average, count, sum, max, min), use AVG(), COUNT(), SUM(), MAX(), MIN()
7. NEVER return generic fallback queries like "SELECT * FROM table LIMIT 100"
8. If the request is ambiguous, make reasonable assumptions based on schema

Examples:
User: "Show all employees with salary greater than 70000"
SQL: SELECT * FROM employees WHERE salary > 70000;

User: "Find average salary by department"
SQL: SELECT department, AVG(salary) AS average_salary FROM employees GROUP BY department;

User: "List employees in Engineering department"
SQL: SELECT * FROM employees WHERE UPPER(department) = UPPER('Engineering');

Now generate the SQL for: "{natural_query}"

SQL Query:"""

        last_exc = None
        for attempt in range(1, max_attempts + 1):
            try:
                # Use low temperature for deterministic SQL generation
                generation_config = genai.types.GenerationConfig(
                    temperature=0.1,
                    top_p=0.95,
                    top_k=40,
                    max_output_tokens=512,
                )
                response = self.gemini_model.generate_content(prompt, generation_config=generation_config)
                sql_query = response.text.strip()
                sql_query = self._clean_sql_query(sql_query)

                # Validate the generated SQL against the user's intent; retry with clarifying prompt if it clearly misses
                valid = self._validate_sql_matches_intent(natural_query, sql_query, schema)
                if valid:
                    logger.info(f"Generated SQL with Gemini: {sql_query}")
                    return sql_query
                else:
                    logger.warning(f"Gemini returned SQL that does not match intent: {sql_query}")
                    # Retry with a more aggressive clarification prompt
                    clarification = f"""\n\nYour previous response was incorrect. It did not implement the user's request properly.

User asked: \"{natural_query}\"

You MUST:
- Include WHERE clauses for any filters mentioned (e.g., salary > 70000)
- Use aggregation functions (AVG, COUNT, SUM, MAX, MIN) if the user asks for averages, counts, totals, etc.
- Use GROUP BY when aggregating by categories
- NEVER return a generic 'SELECT * FROM table LIMIT 100'

Generate the CORRECT SQL now. Return ONLY the SQL statement with no explanation.

SQL Query:"""
                    for clar_try in range(2):
                        try:
                            resp2 = self.gemini_model.generate_content(prompt + clarification, generation_config=generation_config)
                            sql2 = getattr(resp2, 'text', '').strip()
                            sql2 = self._clean_sql_query(sql2)
                            if sql2 and 'NO_SQL_POSSIBLE' not in sql2.upper():
                                if self._validate_sql_matches_intent(natural_query, sql2, schema):
                                    logger.info(f"Generated SQL with Gemini after clarification: {sql2}")
                                    return sql2
                                else:
                                    logger.warning(f"Clarified SQL still invalid: {sql2}")
                            else:
                                logger.warning('Gemini declined SQL generation')
                                break
                        except Exception as e2:
                            logger.debug(f"Clarification attempt {clar_try+1} failed: {e2}")
                            break

                    # if we reach here, clarified attempts failed -> continue to retry/backoff
                    logger.debug('Clarification exhausted; continuing with retry/backoff')

            except Exception as e:
                last_exc = e
                msg = str(e)
                logger.error(f"Error generating SQL with Gemini: {e}")

                # Try to parse retry_delay seconds from the exception message if available
                retry_seconds = None
                try:
                    m = re.search(r"retry in (\d+(?:\.\d+)?)s", msg)
                    if m:
                        retry_seconds = float(m.group(1))
                    else:
                        # Some Gemini errors include 'retry_delay { seconds: 40 }' style
                        m2 = re.search(r"retry_delay\s*\{\s*seconds:\s*(\d+)", msg)
                        if m2:
                            retry_seconds = float(m2.group(1))
                except Exception:
                    retry_seconds = None

                # If we have an explicit retry_seconds, respect it; otherwise use exponential backoff
                if retry_seconds is not None and attempt < max_attempts:
                    wait = retry_seconds
                    logger.info(f"Gemini requested retry after {wait}s; sleeping (attempt {attempt})")
                    import time
                    time.sleep(wait)
                    continue
                elif attempt < max_attempts:
                    wait = backoff_base ** attempt
                    logger.info(f"Retrying Gemini after {wait:.1f}s (attempt {attempt})")
                    import time
                    time.sleep(wait)
                    continue
                else:
                    break

        # If we get here, Gemini failed after retries — fall back to deterministic rules
        logger.warning("Gemini failed after retries — falling back to rule-based SQL generator")
        return self._fallback_sql(natural_query, schema)

    def _fallback_sql(self, natural_query: str, schema: Dict) -> str:
        """Return a deterministic SQL via rule-based heuristics or a safe default SELECT."""
        try:
            rb = self._rule_based_sql(natural_query, schema)
            if rb:
                return rb
        except Exception as e:
            logger.debug(f"Rule-based fallback failed: {e}")

        # As a last resort, return a safe SELECT from a likely table
        q = natural_query.lower()
        tbl_m = re.search(r"(employees|departments|projects|products|students)", q)
        tbl = tbl_m.group(1) if tbl_m else 'employees'
        logger.info(f"Using last-resort fallback SELECT from '{tbl}'")
        return f"SELECT * FROM {tbl} LIMIT 100;"
    
    def _format_schema_for_prompt(self, schema: Dict) -> str:
        """Format schema information for AI prompt"""
        schema_text = ""
        
        for table_name, table_info in schema.items():
            schema_text += f"\nTable: {table_name} (rows: {table_info['row_count']})\n"
            schema_text += "Columns:\n"
            
            for col_name, col_info in table_info['columns'].items():
                schema_text += f"  - {col_name} ({col_info['type']}"
                if col_info['primary_key']:
                    schema_text += ", PRIMARY KEY"
                if col_info['not_null']:
                    schema_text += ", NOT NULL"
                schema_text += ")\n"
                
        return schema_text
    
    def _clean_sql_query(self, sql_query: str) -> str:
        """Clean and validate SQL query"""
        # Remove common markdown formatting
        sql_query = re.sub(r'```sql\n?', '', sql_query)
        sql_query = re.sub(r'```\n?', '', sql_query)
        
        # Remove extra whitespace
        sql_query = ' '.join(sql_query.split())

        # Try to extract the first SQL statement starting with a common SQL keyword
        match = re.search(r"\b(SELECT|INSERT|UPDATE|DELETE|CREATE|ALTER|DROP)\b.*?;", sql_query, re.IGNORECASE)
        if match:
            stmt = match.group(0).strip()
        else:
            # If no semicolon-terminated statement, try to extract from first keyword to end
            match2 = re.search(r"\b(SELECT|INSERT|UPDATE|DELETE|CREATE|ALTER|DROP)\b.*", sql_query, re.IGNORECASE)
            if match2:
                stmt = match2.group(0).strip()
            else:
                # Fallback: return the cleaned string as-is
                stmt = sql_query.strip()

        # Ensure statement ends with a semicolon
        if not stmt.endswith(';'):
            stmt = stmt + ';'

        return stmt

    def _validate_sql_matches_intent(self, natural_query: str, sql_query: str, schema: Dict) -> bool:
        """Basic heuristic validator that checks whether generated SQL likely matches the user's intent.

        This is intentionally conservative: it checks for presence of important column mentions,
        simple comparators for numeric filters, and aggregation keywords when requested.
        It is not a full semantic validator but helps catch obvious mismatches like 'SELECT * FROM ... LIMIT 100;'.
        """
        try:
            nq = natural_query.lower()
            sq = sql_query.lower()

            # Allow CREATE and INSERT statements without validation
            if sq.strip().startswith(('create table', 'create index', 'insert into', 'update ', 'delete from')):
                return True

            # If SQL is the generic fallback pattern, reject
            if re.match(r"select\s+\*\s+from\s+\w+\s+limit\s+\d+;?", sq.strip()):
                return False

            # Check for aggregation intent
            if any(tok in nq for tok in ['average', 'avg', 'count', 'sum', 'maximum', 'minimum', 'max', 'min']):
                if not any(fn in sq for fn in ['avg(', 'count(', 'sum(', 'max(', 'min(']):
                    return False

            # Check for numeric comparison mention
            m_nums = re.findall(r"\b(\d{2,})\b", nq)
            if m_nums:
                # Ensure the number appears in SQL or there's a comparator
                if not any(num in sq for num in m_nums) and not any(op in sq for op in ['>', '<', '>=', '<=']):
                    return False

            # Map schema columns and ensure at least one column referenced in natural query appears in SQL
            mentioned_cols = set()
            for table, info in schema.items():
                for col in info.get('columns', {}).keys():
                    if col.lower() in nq:
                        mentioned_cols.add(col.lower())

            if mentioned_cols:
                if not any(col in sq for col in mentioned_cols):
                    return False

            # If no blocking heuristics matched, accept
            return True
        except Exception:
            return True
    
    def execute_sql_with_metadata(self, sql_query: str) -> QueryResult:
        """Execute SQL query and return detailed results"""
        start_time = datetime.now()
        
        try:
            # Open a short-lived connection for thread-safety
            logger.info(f"[EXECUTE] Using database: {self.db_path}")
            conn = sqlite3.connect(self.db_path, check_same_thread=False)
            try:
                if sql_query.strip().upper().startswith('SELECT'):
                    df = pd.read_sql_query(sql_query, conn)
                    data = df
                else:
                    cur = conn.cursor()
                    cur.execute(sql_query)
                    conn.commit()
                    affected_rows = cur.rowcount

                    # If UPDATE affected 0 rows, attempt a case-insensitive fallback for simple equality WHERE clauses
                    try:
                        if affected_rows == 0 and sql_query.strip().upper().startswith('UPDATE'):
                            # Try to parse SET and WHERE parts
                            upd_match = re.match(r"UPDATE\s+([a-zA-Z_]\w*)\s+SET\s+(.*?)\s+WHERE\s+(.+)",
                                                 sql_query.strip().rstrip(';'), re.IGNORECASE | re.DOTALL)
                            if upd_match:
                                tbl_name = upd_match.group(1)
                                set_clause = upd_match.group(2).strip()
                                where_clause_raw = upd_match.group(3).strip().rstrip(';')

                                # Build case-insensitive WHERE for searching matching rows
                                try:
                                    ci_where = re.sub(r"([a-zA-Z_]\w*)\s*=\s*'([^']*)'",
                                                      lambda m: f"lower({m.group(1)}) = lower('{m.group(2)}')",
                                                      where_clause_raw)
                                    sel_rowids = f"SELECT rowid FROM {tbl_name} WHERE {ci_where}"
                                    logger.debug(f"Case-insensitive fallback: ci_where='{ci_where}', sel_rowids='{sel_rowids}'")
                                    ids_df = pd.read_sql_query(sel_rowids, conn)
                                    logger.debug(f"Case-insensitive fallback: ids_df rows={0 if ids_df is None else len(ids_df)}")
                                    if ids_df is not None and not ids_df.empty:
                                        rowids = ids_df['rowid'].astype(int).tolist()
                                        # Run targeted UPDATE using rowid IN (...) to apply the change
                                        ids_list = ','.join(str(int(r)) for r in rowids)
                                        fallback_update = f"UPDATE {tbl_name} SET {set_clause} WHERE rowid IN ({ids_list})"
                                        try:
                                            cur.execute(fallback_update)
                                            conn.commit()
                                            affected_rows = cur.rowcount
                                            logger.info(f"Applied case-insensitive fallback UPDATE on {tbl_name} for rowids={ids_list}; affected_rows={affected_rows}")
                                            # After fallback update, fetch the updated rows to return
                                            sel_updated = f"SELECT * FROM {tbl_name} WHERE rowid IN ({ids_list})"
                                            df_updated = pd.read_sql_query(sel_updated, conn)
                                            if df_updated is not None:
                                                data = df_updated
                                        except Exception as e:
                                            logger.debug(f"Fallback UPDATE failed: {e}")
                                except Exception as e:
                                    logger.debug(f"Case-insensitive fallback search failed: {e}")
                    except Exception:
                        # Non-fatal
                        pass
                    # Default payload: affected_rows count
                    data = pd.DataFrame({'affected_rows': [affected_rows]})

                    try:
                        sql_norm = sql_query.strip().rstrip(';')

                        # INSERT: try to return the inserted rows using lastrowid or recent rows
                        ins_m = re.match(r"INSERT\s+INTO\s+([a-zA-Z_]\w*)", sql_norm, re.IGNORECASE)
                        if ins_m:
                            tbl = ins_m.group(1)
                            # If lastrowid is available, prefer selecting that row
                            lastrowid = getattr(cur, 'lastrowid', None)
                            if lastrowid and int(lastrowid) > 0:
                                try:
                                    sel = f"SELECT * FROM {tbl} WHERE rowid = {int(lastrowid)}"
                                    df = pd.read_sql_query(sel, conn)
                                    data = df
                                except Exception:
                                    # fallback to recent rows
                                    sel2 = f"SELECT * FROM {tbl} ORDER BY rowid DESC LIMIT {max(affected_rows,1)}"
                                    df = pd.read_sql_query(sel2, conn)
                                    data = df
                            else:
                                # fallback: return most recent N rows
                                sel2 = f"SELECT * FROM {tbl} ORDER BY rowid DESC LIMIT {max(affected_rows,1)}"
                                try:
                                    df = pd.read_sql_query(sel2, conn)
                                    data = df
                                except Exception:
                                    pass
                        else:
                            # UPDATE: try to extract table, SET clause, and WHERE clause to return updated rows
                            upd_m = re.match(r"UPDATE\s+([a-zA-Z_]\w*)\s+SET\s+(.*?)\s+WHERE\s+(.+)", sql_norm, re.IGNORECASE | re.DOTALL)
                            if upd_m:
                                tbl = upd_m.group(1)
                                set_clause = upd_m.group(2).strip()
                                where_clause = upd_m.group(3).strip().rstrip(';')
                                
                                # Extract columns that were SET (updated) to exclude them from WHERE
                                updated_cols = set()
                                for assignment in re.finditer(r"([a-zA-Z_]\w*)\s*=", set_clause):
                                    updated_cols.add(assignment.group(1).lower())
                                
                                # Build a modified WHERE clause excluding updated columns
                                # This ensures we can find rows after the update
                                where_parts = []
                                for condition in re.split(r'\s+AND\s+', where_clause, flags=re.IGNORECASE):
                                    # Check if this condition involves an updated column
                                    col_match = re.match(r"([a-zA-Z_]\w*)\s*=", condition.strip())
                                    if col_match:
                                        col_name = col_match.group(1).lower()
                                        # Only keep conditions for columns that were NOT updated
                                        if col_name not in updated_cols:
                                            where_parts.append(condition.strip())
                                    else:
                                        # Keep non-equality conditions
                                        where_parts.append(condition.strip())
                                
                                try:
                                    # Try to select with modified WHERE (excluding updated columns)
                                    if where_parts:
                                        modified_where = ' AND '.join(where_parts)
                                        sel = f"SELECT * FROM {tbl} WHERE {modified_where}"
                                        df = pd.read_sql_query(sel, conn)
                                        
                                        # If still empty, try case-insensitive
                                        if df is None or df.empty:
                                            ci_where = re.sub(r"([a-zA-Z_]\w*)\s*=\s*'([^']*)'",
                                                              lambda m: f"lower({m.group(1)}) = lower('{m.group(2)}')",
                                                              modified_where)
                                            sel_ci = f"SELECT * FROM {tbl} WHERE {ci_where}"
                                            df_ci = pd.read_sql_query(sel_ci, conn)
                                            if df_ci is not None and not df_ci.empty:
                                                df = df_ci
                                    else:
                                        # No WHERE parts left (all were updated columns)
                                        # Return recent rows that match the SET values
                                        set_conditions = []
                                        for assignment in re.finditer(r"([a-zA-Z_]\w*)\s*=\s*(.+?)(?:,|$)", set_clause):
                                            col = assignment.group(1).strip()
                                            val = assignment.group(2).strip()
                                            set_conditions.append(f"{col} = {val}")
                                        
                                        if set_conditions:
                                            set_where = ' AND '.join(set_conditions)
                                            sel = f"SELECT * FROM {tbl} WHERE {set_where}"
                                            df = pd.read_sql_query(sel, conn)
                                        else:
                                            # Last resort: recent rows
                                            sel = f"SELECT * FROM {tbl} ORDER BY rowid DESC LIMIT {max(affected_rows, 10)}"
                                            df = pd.read_sql_query(sel, conn)
                                    
                                    data = df
                                except Exception as e:
                                    logger.debug(f"UPDATE post-processing with modified WHERE failed: {e}")
                                    # Fallback: return recent rows from table
                                    try:
                                        sel2 = f"SELECT * FROM {tbl} ORDER BY rowid DESC LIMIT 100"
                                        df = pd.read_sql_query(sel2, conn)
                                        data = df
                                    except Exception:
                                        pass
                            else:
                                # DELETE or other DML: try to detect table to return remaining rows
                                del_m = re.match(r"DELETE\s+FROM\s+([a-zA-Z_]\w*)\s*(WHERE\s+.*)?", sql_norm, re.IGNORECASE | re.DOTALL)
                                if del_m:
                                    tbl = del_m.group(1)
                                    try:
                                        sel = f"SELECT * FROM {tbl} ORDER BY rowid DESC LIMIT 100"
                                        df = pd.read_sql_query(sel, conn)
                                        data = df
                                    except Exception:
                                        pass

                    except Exception as e:
                        logger.debug(f"DML post-processing failed: {e}")
                        # keep affected_rows fallback
                        pass
            finally:
                conn.close()
            
            execution_time = (datetime.now() - start_time).total_seconds()
            
            # Ensure we always return a QueryResult instance
            if isinstance(data, QueryResult):
                return data

            return QueryResult(
                success=True,
                data=data,
                sql_query=sql_query,
                execution_time=execution_time
            )
            
        except Exception as e:
            execution_time = (datetime.now() - start_time).total_seconds()
            error_msg = str(e)
            
            logger.error(f"SQL execution error: {error_msg}")
            
            return QueryResult(
                success=False,
                data=None,
                sql_query=sql_query,
                execution_time=execution_time,
                error_message=error_msg
            )
    
    def create_visualization(self, result: QueryResult, chart_type: str = None) -> str:
        """Create visualization for query results"""
        if not result.success or result.data is None or result.data.empty:
            logger.warning("Cannot create visualization: No valid data")
            return None
            
        title = f"Results for: {result.sql_query[:50]}..."
        viz_path = self.visualization_engine.create_visualization(
            result.data, chart_type, title
        )
        
        return viz_path
        
    def get_available_ollama_models(self) -> List[str]:
        """Get list of available Ollama models"""
        try:
            response = requests.get(f"{self.ollama_base_url}/api/tags", timeout=5)
            if response.status_code == 200:
                models = response.json().get('models', [])
                return [model.get('name', '') for model in models]
            return []
        except Exception as e:
            logger.error(f"Failed to get Ollama models: {e}")
            return []
    
    def get_model_status(self) -> Dict[str, Any]:
        """Get current model configuration status"""
        status = {
            'similarity_model': {
                'name': 'paraphrase-multilingual-MiniLM-L12-v2',
                'type': 'offline',
                'purpose': 'Query similarity matching (Telugu/Hindi/English)',
                'status': 'active'
            }
        }
        
        if self.use_ollama:
            status['llm_model'] = {
                'name': self.ollama_model,
                'type': 'offline (Ollama)',
                'purpose': 'Natural language to SQL conversion',
                'status': 'active' if self._test_ollama_connection() else 'error',
                'base_url': self.ollama_base_url
            }
        elif self.gemini_model:
            status['llm_model'] = {
                'name': 'models/gemini-2.0-flash',
                'type': 'online (Google AI)',
                'purpose': 'Natural language to SQL conversion', 
                'status': 'active'
            }
        else:
            status['llm_model'] = {
                'name': 'None',
                'type': 'not configured',
                'purpose': 'Natural language to SQL conversion',
                'status': 'error'
            }
            
        return status
        
    def create_plotly_visualization(self, result: QueryResult, chart_type: str = None) -> Dict:
        """Create interactive Plotly visualization for web interface"""
        if not result.success or result.data is None or result.data.empty:
            return None
            
        title = f"Query Results: {result.sql_query[:50]}..."
        return self.visualization_engine.create_plotly_visualization(
            result.data, chart_type, title
        )
        
    def export_data(self, data: List[Dict], format: str, filename: str) -> bytes:
        """Export query results to various formats"""
        if not data:
            raise ValueError("No data to export")
            
        df = pd.DataFrame(data)
        
        if format.lower() == 'csv':
            output = io.StringIO()
            df.to_csv(output, index=False)
            return output.getvalue().encode('utf-8')
            
        elif format.lower() == 'excel':
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Query Results', index=False)
                
                # Add formatting
                workbook = writer.book
                worksheet = writer.sheets['Query Results']
                
                # Header formatting
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for col_num, col_name in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                    
            output.seek(0)
            return output.getvalue()
            
        elif format.lower() == 'json':
            return json.dumps(data, indent=2, default=str).encode('utf-8')
            
        else:
            raise ValueError(f"Unsupported export format: {format}")
    
    def process_advanced_query(self, natural_query: str, visualize: bool = False, 
                             chart_type: str = None, use_plotly: bool = False) -> Dict:
        """Main method to process natural language queries"""
        logger.info(f"Processing query: {natural_query}")
        
        # Step 1: Check similarity index for cached queries (if enabled)
        sql_query = None
        if self.similarity_index is not None:
            try:
                similar_queries = self.similarity_index.find_similar(natural_query)
                if similar_queries:
                    logger.info(f"Found {len(similar_queries)} similar queries")
                    best_match = similar_queries[0]
                    if best_match.get('similarity', 0) > 0.9:  # Very high similarity
                        logger.info("Using cached query result")
                        sql_query = best_match['query']['sql_query']
            except Exception as e:
                logger.warning(f"Similarity check failed: {e}")
        
        # Step 2: Generate SQL if not found in cache
        if not sql_query:
            schema = self.get_comprehensive_schema()
            sql_query = self.generate_sql_with_ai(natural_query, schema)
            
            if not sql_query:
                return {
                    'success': False,
                    'error': 'Failed to generate SQL query',
                    'natural_query': natural_query
                }
        
        # Step 3: Execute SQL query
        raw_result = self.execute_sql_with_metadata(sql_query)
        if isinstance(raw_result, QueryResult):
            result = raw_result
        else:
            if isinstance(raw_result, pd.DataFrame):
                result = QueryResult(success=True, data=raw_result, sql_query=sql_query, execution_time=0.0)
            else:
                result = QueryResult(success=False, data=None, sql_query=sql_query, execution_time=0.0,
                                     error_message='Unexpected result type from execute_sql_with_metadata')
        
        # Step 4: Add to similarity index if successful and enabled
        if result.success and self.similarity_index is not None and sql_query:
            try:
                metadata = {
                    'timestamp': datetime.now().isoformat(),
                    'row_count': len(result.data) if result.data is not None else 0,
                    'columns': list(result.data.columns) if result.data is not None else []
                }
                self.similarity_index.add_query(natural_query, sql_query, metadata)
            except Exception as e:
                logger.warning(f"Failed to add to similarity index: {e}")
        
        # Step 5: Create visualization if requested
        viz_path = None
        plotly_data = None
        
        if visualize and result.success:
            if use_plotly:
                plotly_data = self.create_plotly_visualization(result, chart_type)
            else:
                viz_path = self.create_visualization(result, chart_type)
                result.visualization_path = viz_path
        
        # Step 6: Format response
        response = {
            'success': result.success,
            'natural_query': natural_query,
            'sql_query': result.sql_query,
            'execution_time': result.execution_time,
            'data': result.data.to_dict('records') if result.data is not None else None,
            'row_count': len(result.data) if result.data is not None else 0,
            'visualization_path': viz_path,
            'visualization_data': plotly_data
        }
        
        if not result.success:
            response['error'] = result.error_message
        
        logger.info(f"Query processed successfully: {result.success}")
        return response
    
    def __del__(self):
        """Clean up database connection"""
        if hasattr(self, 'connection'):
            self.connection.close()

def create_enhanced_converter():
    """Factory function to create converter instance"""
    return AdvancedTextToSQLConverter()

# Flask Web Application
app = Flask(__name__)
CORS(app)

# Global converter instance
converter = None

# Database configuration
DEFAULT_DB_PATH = "data/advanced_nlsql.db"
UPLOAD_FOLDER = "data/uploads"
current_db_path = DEFAULT_DB_PATH
current_db_name = "Default Database"

# Ensure upload folder exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def get_uploaded_databases():
    """Get list of uploaded SQLite databases"""
    databases = []
    if os.path.exists(UPLOAD_FOLDER):
        for f in os.listdir(UPLOAD_FOLDER):
            if f.endswith('.db') or f.endswith('.sqlite') or f.endswith('.sqlite3'):
                db_path = os.path.join(UPLOAD_FOLDER, f)
                databases.append({
                    'name': f,
                    'path': db_path,
                    'size': os.path.getsize(db_path),
                    'modified': datetime.fromtimestamp(os.path.getmtime(db_path)).isoformat()
                })
    return databases

@app.route('/')
def index():
    """Main web interface"""
    return render_template('index.html')

@app.route('/api/similarity', methods=['POST'])
def api_similarity():
    """API endpoint for checking query similarity"""
    try:
        data = request.get_json()
        natural_query = data.get('query', '').strip()
        
        if not natural_query:
            return jsonify({'success': False, 'error': 'Query cannot be empty'})
        
        # If similarity is disabled, return that information
        if converter.similarity_index is None:
            return jsonify({'success': True, 'found_similar': False, 'message': 'Similarity disabled on server'})

        # Check similarity index
        similar_queries = converter.similarity_index.find_similar(natural_query)

        if similar_queries:
            best_match = similar_queries[0]
            return jsonify({
                'success': True,
                'found_similar': True,
                'similarity_score': best_match['similarity'],
                'similar_query': best_match['query']['natural_query'],
                'cached_sql': best_match['query']['sql_query'],
                'use_cached': best_match['similarity'] > 0.9
            })
        else:
            return jsonify({
                'success': True,
                'found_similar': False,
                'message': 'No similar queries found in cache'
            })
            
    except Exception as e:
        logger.error(f"API similarity error: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/similarity-check', methods=['POST'])
def api_similarity_check():
    """Enhanced similarity check API endpoint"""
    try:
        data = request.get_json()
        natural_query = data.get('query', '').strip()
        
        if not natural_query:
            return jsonify({'success': False, 'error': 'Query cannot be empty'})
        
        # If similarity is disabled, return quickly
        if converter.similarity_index is None:
            return jsonify({'success': True, 'similar_queries': [], 'count': 0, 'message': 'Similarity disabled on server'})

        # Check similarity index with 70% threshold (good balance)
        # Lower threshold = more matches but less accurate
        # Higher threshold = fewer matches but more accurate
        similar_queries = converter.similarity_index.find_similar(natural_query, threshold=0.70)
        
        logger.info(f"Similarity check for: '{natural_query}' - Found {len(similar_queries)} matches")

        return jsonify({
            'success': True,
            'similar_queries': similar_queries,
            'count': len(similar_queries)
        })
        
    except Exception as e:
        logger.error(f"API similarity-check error: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/swap-entities', methods=['POST'])
def api_swap_entities():
    """
    Smart entity swapping endpoint
    Takes original query, new query, and original SQL
    Returns adapted SQL with entities swapped
    """
    try:
        data = request.get_json()
        original_query = data.get('original_query', '').strip()
        new_query = data.get('new_query', '').strip()
        original_sql = data.get('original_sql', '').strip()
        
        if not all([original_query, new_query, original_sql]):
            return jsonify({
                'success': False, 
                'error': 'Missing required fields: original_query, new_query, original_sql'
            })
        
        # If similarity is disabled, return original SQL
        if converter.similarity_index is None:
            return jsonify({
                'success': True,
                'adapted_sql': original_sql,
                'swapped': False,
                'structural_change': False,
                'message': 'Similarity disabled - returning original SQL'
            })
        
        # Perform entity swapping - now returns a dict
        result = converter.similarity_index.swap_entities(
            original_query, 
            new_query, 
            original_sql
        )
        
        return jsonify({
            'success': True,
            'adapted_sql': result['adapted_sql'],
            'original_sql': original_sql,
            'swapped': result['swapped'],
            'structural_change': result['structural_change'],
            'message': result['message']
        })
        
    except Exception as e:
        logger.error(f"API swap-entities error: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/ollama-models', methods=['GET'])
def api_ollama_models():
    """Get available Ollama models"""
    try:
        ollama_models = converter.get_available_ollama_models()
        return jsonify(ollama_models)
        
    except Exception as e:
        logger.error(f"API ollama-models error: {e}")
        return jsonify([])

@app.route('/api/models', methods=['GET'])
def api_models():
    """API endpoint for getting available models"""
    try:
        model_status = converter.get_model_status()
        ollama_models = converter.get_available_ollama_models() if converter.use_ollama else []
        
        return jsonify({
            'success': True,
            'current_config': model_status,
            'ollama_available': len(ollama_models) > 0,
            'ollama_models': ollama_models,
            'gemini_available': bool(os.getenv('GEMINI_API_KEY') and os.getenv('GEMINI_API_KEY') != 'your_gemini_api_key_here')
        })
        
    except Exception as e:
        logger.error(f"API models error: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/generate-sql', methods=['POST'])
def api_generate_sql():
    """API endpoint for generating SQL with selected model"""
    try:
        data = request.get_json()
        natural_query = data.get('query', '').strip()
        model = data.get('model', 'models/gemini-2.0-flash')
        use_ollama = data.get('use_ollama', False)
        
        if not natural_query:
            return jsonify({'success': False, 'error': 'Query cannot be empty'})
        
        # Temporarily override model settings for this request
        original_use_ollama = converter.use_ollama
        original_ollama_model = converter.ollama_model
        
        # If user requested Ollama or the chosen model is not a Gemini model
        if use_ollama or ('gemini' not in model):
            converter.use_ollama = True
            converter.ollama_model = model
        else:
            converter.use_ollama = False
        
        # Generate SQL
        schema = converter.get_comprehensive_schema()
        sql_query = converter.generate_sql_with_ai(natural_query, schema)
        
        # Restore original settings
        converter.use_ollama = original_use_ollama
        converter.ollama_model = original_ollama_model
        
        if sql_query:
            # AUTO-LEARNING: Save LLM-generated query to FAISS for future users
            try:
                converter.similarity_index.add_query(
                    query=natural_query,
                    sql=sql_query,
                    result_metadata={
                        'source': 'llm_generated',
                        'model': model,
                        'generated_at': datetime.now().isoformat()
                    }
                )
                logger.info(f"[AUTO-LEARN] Saved new query to FAISS: {natural_query[:50]}...")
            except Exception as e:
                logger.warning(f"Failed to auto-learn query: {e}")
            
            return jsonify({
                'success': True,
                'sql_query': sql_query,
                'model_used': model
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Failed to generate SQL query'
            })
            
    except Exception as e:
        logger.error(f"API generate-sql error: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/schema', methods=['GET'])
def api_schema():
    """API endpoint for getting database schema"""
    try:
        schema = converter.get_comprehensive_schema()
        return jsonify(schema)
    except Exception as e:
        logger.error(f"API schema error: {e}")
        return jsonify({})

@app.route('/api/export', methods=['POST'])
def api_export():
    """API endpoint for exporting data"""
    try:
        data = request.get_json()
        export_data = data.get('data', [])
        export_format = data.get('format', 'csv')
        filename = data.get('filename', 'export')
        
        if not export_data:
            return jsonify({'success': False, 'error': 'No data to export'})
        
        exported_data = converter.export_data(export_data, export_format, filename)
        
        # Determine content type and file extension
        content_types = {
            'csv': 'text/csv',
            'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'json': 'application/json'
        }
        
        extensions = {
            'csv': 'csv',
            'excel': 'xlsx', 
            'json': 'json'
        }
        
        return send_file(
            io.BytesIO(exported_data),
            mimetype=content_types.get(export_format, 'text/plain'),
            as_attachment=True,
            download_name=f"{filename}.{extensions.get(export_format, 'txt')}"
        )
        
    except Exception as e:
        logger.error(f"API export error: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/execute-sql', methods=['POST'])
def api_execute_sql():
    """API endpoint for executing SQL query directly"""
    try:
        data = request.get_json()
        sql_query = data.get('sql_query', '').strip()
        natural_query = data.get('natural_query', '')
        visualize = data.get('visualize', False)
        chart_type = data.get('chart_type', None)
        
        if not sql_query:
            return jsonify({'success': False, 'error': 'SQL query cannot be empty'})
        
        # Execute SQL query
        raw_result = converter.execute_sql_with_metadata(sql_query)
        # Defensive: ensure we have a QueryResult
        if isinstance(raw_result, QueryResult):
            result = raw_result
        else:
            # If a DataFrame (or other) was returned, wrap it
            if hasattr(raw_result, 'to_dict') or isinstance(raw_result, pd.DataFrame):
                df = raw_result if isinstance(raw_result, pd.DataFrame) else None
                result = QueryResult(success=True, data=df, sql_query=sql_query, execution_time=0.0)
            else:
                # Unknown type -> create a failed QueryResult
                result = QueryResult(success=False, data=None, sql_query=sql_query, execution_time=0.0,
                                     error_message='Unexpected result type from execute_sql_with_metadata')
        
        # Add to similarity index if successful (only if similarity is enabled)
        if result.success and natural_query:
            metadata = {
                'timestamp': datetime.now().isoformat(),
                'row_count': len(result.data) if result.data is not None else 0,
                'columns': list(result.data.columns) if result.data is not None else []
            }
            try:
                if getattr(converter, 'similarity_index', None) is not None:
                    converter.similarity_index.add_query(natural_query, sql_query, metadata)
                else:
                    logger.debug("Similarity index not available; skipping add_query")
            except Exception as e:
                # Do not let similarity failures break the execute flow
                logger.warning(f"Failed to add query to similarity index: {e}")
        
        # Create visualization if requested
        plotly_data = None
        if visualize and result.success and result.data is not None:
            # create_plotly_visualization expects a QueryResult object
            plotly_data = converter.create_plotly_visualization(result, chart_type)
        
        # Format response
        response = {
            'success': result.success,
            'sql_query': result.sql_query,
            'execution_time': result.execution_time,
            'data': result.data.to_dict('records') if result.data is not None else None,
            'row_count': len(result.data) if result.data is not None else 0,
            'visualization_data': plotly_data,
            'database_used': converter.db_path  # Show which database was used
        }
        
        if not result.success:
            response['error'] = result.error_message
        
        return jsonify(response)
        
    except Exception as e:
        logger.error(f"API execute-sql error: {e}")
        return jsonify({'success': False, 'error': str(e)})

# ============= DATABASE MANAGEMENT ENDPOINTS =============

@app.route('/api/db/current', methods=['GET'])
def api_db_current():
    """Get current database information"""
    global current_db_path, current_db_name
    try:
        # Get tables from current database using a new connection
        tables = []
        try:
            conn = sqlite3.connect(current_db_path, check_same_thread=False)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'")
            tables = [row[0] for row in cursor.fetchall()]
            conn.close()
        except Exception as e:
            logger.warning(f"Could not read tables from current db: {e}")
        
        return jsonify({
            'success': True,
            'current_db': {
                'name': current_db_name,
                'path': current_db_path,
                'tables': tables,
                'table_count': len(tables)
            },
            'uploaded_databases': get_uploaded_databases()
        })
    except Exception as e:
        logger.error(f"API db/current error: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/db/connect-path', methods=['POST'])
def api_db_connect_path():
    """Connect directly to a database file by path (changes will be made to original file)"""
    global converter, current_db_path, current_db_name
    try:
        data = request.get_json()
        db_path = data.get('path', '').strip()
        
        if not db_path:
            return jsonify({'success': False, 'error': 'Database path not provided'})
        
        # Normalize path
        db_path = os.path.normpath(db_path)
        
        # Check if file exists
        if not os.path.exists(db_path):
            return jsonify({'success': False, 'error': f'File not found: {db_path}'})
        
        # Validate it's a valid SQLite database
        try:
            test_conn = sqlite3.connect(db_path)
            cursor = test_conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [row[0] for row in cursor.fetchall()]
            test_conn.close()
            
            if not tables:
                return jsonify({
                    'success': False,
                    'error': 'The file contains no tables. Please provide a valid SQLite database.'
                })
                
        except sqlite3.DatabaseError as e:
            return jsonify({
                'success': False,
                'error': f'Invalid SQLite database file: {str(e)}'
            })
        
        # Switch to the database - connect directly to the original file
        current_db_path = db_path
        current_db_name = os.path.basename(db_path) + " (Direct)"
        converter.db_path = db_path
        
        # Get schema info
        schema = converter.get_comprehensive_schema()
        
        logger.info(f"Connected directly to database: {db_path}")
        
        return jsonify({
            'success': True,
            'message': f'Connected directly to "{os.path.basename(db_path)}" - changes will be saved to original file',
            'database': {
                'name': current_db_name,
                'path': db_path,
                'tables': list(schema.keys()),
                'table_count': len(schema)
            }
        })
        
    except Exception as e:
        logger.error(f"API db/connect-path error: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/db/upload', methods=['POST'])
def api_db_upload():
    """Upload a SQLite database file"""
    global converter, current_db_path, current_db_name
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'})
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'})
        
        # Check file extension
        allowed_extensions = {'.db', '.sqlite', '.sqlite3'}
        file_ext = os.path.splitext(file.filename)[1].lower()
        
        if file_ext not in allowed_extensions:
            return jsonify({
                'success': False, 
                'error': f'Invalid file type. Allowed: {", ".join(allowed_extensions)}'
            })
        
        # Save the file
        filename = file.filename.replace(' ', '_')
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)
        
        # Validate it's a valid SQLite database
        try:
            test_conn = sqlite3.connect(filepath)
            cursor = test_conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [row[0] for row in cursor.fetchall()]
            test_conn.close()
            
            if not tables:
                os.remove(filepath)
                return jsonify({
                    'success': False,
                    'error': 'The uploaded file contains no tables. Please upload a valid SQLite database.'
                })
                
        except sqlite3.DatabaseError as e:
            os.remove(filepath)
            return jsonify({
                'success': False,
                'error': f'Invalid SQLite database file: {str(e)}'
            })
        
        # Switch to the new database - just update db_path
        current_db_path = filepath
        current_db_name = filename
        converter.db_path = filepath
        
        # Get schema info
        schema = converter.get_comprehensive_schema()
        
        logger.info(f"Switched to uploaded database: {filename}")
        
        return jsonify({
            'success': True,
            'message': f'Database "{filename}" uploaded and activated',
            'database': {
                'name': filename,
                'path': filepath,
                'tables': list(schema.keys()),
                'table_count': len(schema)
            }
        })
        
    except Exception as e:
        logger.error(f"API db/upload error: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/db/switch', methods=['POST'])
def api_db_switch():
    """Switch to a different database"""
    global converter, current_db_path, current_db_name
    try:
        data = request.get_json()
        db_path = data.get('path', '')
        
        if not db_path:
            return jsonify({'success': False, 'error': 'Database path not provided'})
        
        # Check if it's the default database
        if db_path == 'default':
            db_path = DEFAULT_DB_PATH
            db_name = 'Default Database'
        else:
            if not os.path.exists(db_path):
                return jsonify({'success': False, 'error': 'Database file not found'})
            db_name = os.path.basename(db_path)
        
        # Switch database - just update db_path
        converter.db_path = db_path
        current_db_path = db_path
        current_db_name = db_name
        
        # Get schema info
        schema = converter.get_comprehensive_schema()
        
        logger.info(f"Switched to database: {db_name}")
        
        return jsonify({
            'success': True,
            'message': f'Switched to "{db_name}"',
            'database': {
                'name': db_name,
                'path': db_path,
                'tables': list(schema.keys()),
                'table_count': len(schema)
            }
        })
        
    except Exception as e:
        logger.error(f"API db/switch error: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/db/delete', methods=['POST'])
def api_db_delete():
    """Delete an uploaded database"""
    global converter, current_db_path, current_db_name
    try:
        data = request.get_json()
        db_path = data.get('path', '')
        
        if not db_path:
            return jsonify({'success': False, 'error': 'Database path not provided'})
        
        # Don't allow deleting the default database
        if db_path == DEFAULT_DB_PATH:
            return jsonify({'success': False, 'error': 'Cannot delete the default database'})
        
        # Check if file exists in uploads folder
        if not db_path.startswith(UPLOAD_FOLDER):
            return jsonify({'success': False, 'error': 'Can only delete uploaded databases'})
        
        if not os.path.exists(db_path):
            return jsonify({'success': False, 'error': 'Database file not found'})
        
        # If this is the current database, switch to default first
        if db_path == current_db_path:
            converter.db_path = DEFAULT_DB_PATH
            current_db_path = DEFAULT_DB_PATH
            current_db_name = 'Default Database'
        
        # Delete the file
        os.remove(db_path)
        
        logger.info(f"Deleted database: {db_path}")
        
        return jsonify({
            'success': True,
            'message': 'Database deleted successfully'
        })
        
    except Exception as e:
        logger.error(f"API db/delete error: {e}")
        return jsonify({'success': False, 'error': str(e)})

def run_web_interface():
    """Run the Flask web application"""
    global converter
    converter = AdvancedTextToSQLConverter()
    
    print("🌐 Starting web interface...")
    print("📊 Open your browser and go to: http://localhost:5000")
    print("🔄 Press Ctrl+C to stop the server")
    
    try:
        app.run(host='0.0.0.0', port=5000, debug=False)
    except KeyboardInterrupt:
        print("\n👋 Web server stopped!")

def run_cli_interface():
    """Run the command line interface"""
    print("🚀 Advanced NL-to-SQL System - CLI Mode")
    print("=" * 50)
    
    # Initialize the system
    converter = AdvancedTextToSQLConverter()
    
    # Show model status
    model_status = converter.get_model_status()
    print("\n🤖 Model Configuration:")
    print("-" * 30)
    
    for model_type, info in model_status.items():
        status_icon = "✅" if info['status'] == 'active' else "❌"
        model_name = info['name']
        model_type_str = info['type']
        
        print(f"{status_icon} {info['purpose']}:")
        print(f"   Model: {model_name} ({model_type_str})")
        
        if 'base_url' in info:
            print(f"   URL: {info['base_url']}")
    
    # Show available Ollama models if Ollama is being used
    if converter.use_ollama:
        available_models = converter.get_available_ollama_models()
        if available_models:
            print(f"\n📋 Available Ollama Models:")
            for model in available_models:
                icon = "✅" if model == converter.ollama_model else "▪️"
                print(f"   {icon} {model}")
    
    # Test queries
    test_queries = [
        "Show all employees with their departments",
        "Find average salary by department", 
        "Show employees hired in 2023",
        "List all projects with their status",
        "Show departments with budget greater than 200000",
        "Find the highest paid employee in each department"
    ]
    
    print("\n📋 Sample queries you can try:")
    for i, query in enumerate(test_queries, 1):
        print(f"{i}. {query}")
    
    print("\n💡 Interactive Mode - Type your natural language queries:")
    print("Commands: 'quit' to exit, 'schema' to view database schema, 'models' to check model status")
    print("-" * 50)
    
    while True:
        try:
            query = input("\n💬 Your query: ").strip()
            
            if query.lower() in ['quit', 'exit', 'q']:
                print("👋 Goodbye!")
                break
                
            if query.lower() == 'models':
                model_status = converter.get_model_status()
                print("\n🤖 Current Model Status:")
                for model_type, info in model_status.items():
                    status_icon = "✅" if info['status'] == 'active' else "❌"
                    print(f"{status_icon} {info['purpose']}: {info['name']} ({info['type']})")
                continue
                
            if query.lower() == 'schema':
                schema = converter.get_comprehensive_schema()
                print("\n📊 Database Schema:")
                for table, info in schema.items():
                    print(f"\n📋 Table: {table} ({info['row_count']} rows)")
                    for col, col_info in info['columns'].items():
                        print(f"   - {col} ({col_info['type']})")
                continue
                
            if not query:
                continue
            
            # Check if user wants visualization
            visualize = input("📊 Create visualization? (y/n): ").lower().startswith('y')
            chart_type = None
            
            if visualize:
                chart_type = input("📈 Chart type (bar/pie/line/histogram/scatter/auto): ").strip()
                if chart_type == 'auto' or not chart_type:
                    chart_type = None
            
            print("\n⏳ Processing...")
            result = converter.process_advanced_query(query, visualize, chart_type)
            
            if result['success']:
                print(f"\n✅ Success!")
                print(f"🔧 SQL: {result['sql_query']}")
                print(f"⏱️  Execution time: {result['execution_time']:.3f}s")
                
                if result.get('data') and len(result['data']) > 0:
                    print(f"📊 Rows: {result['row_count']}")
                    # Show first few rows for SELECT queries
                    if result['row_count'] <= 10:
                        df = pd.DataFrame(result['data'])
                        print("📋 Data:")
                        print(df.to_string(index=False))
                    else:
                        print(f"📋 First 3 rows: {result['data'][:3]}")
                        
                if result.get('visualization_path'):
                    print(f"📈 Chart saved: {result['visualization_path']}")
                    
            else:
                print(f"\n❌ Error: {result.get('error')}")
                
        except KeyboardInterrupt:
            print("\n\n👋 Goodbye!")
            break
        except Exception as e:
            print(f"❌ Unexpected error: {e}")

def main():
    """Main function with interface selection"""
    import sys
    
    # Check for command line arguments
    if len(sys.argv) > 1 and sys.argv[1] == '--web':
        run_web_interface()
        return
    
    print("🚀 Advanced NL-to-SQL System")
    print("=" * 40)
    print("Select interface mode:")
    print("1. 💻 Command Line Interface")
    print("2. 🌐 Web Interface")
    print("3. 🧪 Quick Demo")
    
    while True:
        choice = input("\nEnter your choice (1-3): ").strip()
        
        if choice == '1':
            run_cli_interface()
            break
        elif choice == '2':
            run_web_interface()
            break
        elif choice == '3':
            # Quick demo mode (original main function)
            print("\n🧪 Running Quick Demo...")
            print("=" * 30)
            
            # Initialize the system
            converter = AdvancedTextToSQLConverter()
            
            # Demo queries
            demo_queries = [
                ("Show all employees in Engineering department", True, "bar"),
                ("Find average salary by department", True, "pie"),
                ("List all projects with status 'In Progress'", False, None),
                ("Show departments with budget greater than 200000", True, "bar")
            ]
            
            print("Running demo queries with visualizations:")
            
            for i, (query, viz, chart) in enumerate(demo_queries, 1):
                print(f"\n{i}. Query: {query}")
                result = converter.process_advanced_query(query, viz, chart)
                
                if result['success']:
                    print(f"   ✅ SQL: {result['sql_query'][:60]}...")
                    if result.get('data'):
                        print(f"   📊 Rows returned: {result['row_count']}")
                    if result.get('visualization_path'):
                        print(f"   📈 Visualization: {result['visualization_path']}")
                else:
                    print(f"   ❌ Error: {result.get('error', 'Unknown error')}")
            
            # Show final schema
            schema = converter.get_comprehensive_schema()
            print(f"\n📋 Database contains {len(schema)} tables:")
            for table, info in schema.items():
                print(f"   - {table}: {info['row_count']} rows")
                
            print("\n🎉 Demo completed! Check 'visualizations' folder for charts.")
            break
        else:
            print("Invalid choice. Please enter 1, 2, or 3.")

if __name__ == "__main__":
    main()