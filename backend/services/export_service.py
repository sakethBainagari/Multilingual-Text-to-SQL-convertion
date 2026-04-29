"""Export query results to CSV / Excel / JSON."""

import io
import json
import logging
import re
from typing import Dict, List

import pandas as pd
from openpyxl.styles import Font, PatternFill

logger = logging.getLogger(__name__)


class ExportService:
    """Convert query result data into downloadable file bytes."""

    @staticmethod
    def sanitize_filename(name: str) -> str:
        """Strip unsafe characters from filenames."""
        return re.sub(r"[^\w\-.]", "_", name)[:100]

    def export(self, data: List[Dict], fmt: str, filename: str = "export") -> bytes:
        """Return *data* serialised in *fmt* (``csv``, ``excel``, or ``json``)."""
        if not data:
            raise ValueError("No data to export")

        filename = self.sanitize_filename(filename)
        df = pd.DataFrame(data)

        if fmt.lower() == "csv":
            return self._to_csv(df)
        if fmt.lower() == "excel":
            return self._to_excel(df)
        if fmt.lower() == "json":
            return self._to_json(data)
        raise ValueError(f"Unsupported export format: {fmt}")

    # ------------------------------------------------------------------

    @staticmethod
    def _to_csv(df: pd.DataFrame) -> bytes:
        buf = io.StringIO()
        df.to_csv(buf, index=False)
        return buf.getvalue().encode("utf-8")

    @staticmethod
    def _to_excel(df: pd.DataFrame) -> bytes:
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Query Results", index=False)
            wb = writer.book
            ws = writer.sheets["Query Results"]

            hdr_font = Font(bold=True, color="FFFFFF")
            hdr_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            for col_num, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = hdr_font
                cell.fill = hdr_fill

            # Auto-width
            for column_cells in ws.columns:
                letter = column_cells[0].column_letter
                max_len = max(len(str(c.value or "")) for c in column_cells)
                ws.column_dimensions[letter].width = min(max_len + 2, 50)

            # Summary row
            summary_row = len(df) + 2
            ws.cell(row=summary_row, column=1, value=f"Total rows: {len(df)}").font = Font(bold=True)

        buf.seek(0)
        return buf.getvalue()

    @staticmethod
    def _to_json(data: List[Dict]) -> bytes:
        return json.dumps(data, indent=2, default=str).encode("utf-8")

    # ------------------------------------------------------------------
    # Content-type / extension helpers
    # ------------------------------------------------------------------

    CONTENT_TYPES = {
        "csv": "text/csv",
        "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "json": "application/json",
    }
    EXTENSIONS = {"csv": "csv", "excel": "xlsx", "json": "json"}
